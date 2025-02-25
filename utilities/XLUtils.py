import openpyxl

def get_row_count(file, sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.max_row

def get_col_count(file, sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.max_col

def get_cell_value(file,sheetname,row_num,col_num):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(row=row_num, column=col_num).value

def set_cell_value(file,sheetname,row_num,col_num, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(row=row_num, column=col_num).value= data
    workbook.save(file)